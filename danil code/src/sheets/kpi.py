from openpyxl.utils import get_column_letter

from src.styles import (
    BG, NUM, PCT, PCT2,
    make_fill, make_font, make_border, make_align, set_col_widths, data_cell,
)


def build(wb, cfg: dict) -> None:
    ws = wb.create_sheet("KPI")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A": 32, "B": 22, "C": 18, "D": 18})

    n = cfg["years"]

    ws.merge_cells("A1:D1")
    ws["A1"].value     = "КЛЮЧЕВЫЕ ПОКАЗАТЕЛИ ЭФФЕКТИВНОСТИ"
    ws["A1"].fill      = make_fill(BG["navy"])
    ws["A1"].font      = make_font(bold=True, size=12, color="FFFFFF")
    ws["A1"].alignment = make_align("center")
    ws.row_dimensions[1].height = 28

    for col_i, text in enumerate(
        ["Показатель", "Описание",
         f"База ({cfg['base_year']})", f"Итог ({cfg['base_year'] + n})"],
        start=1,
    ):
        c = ws.cell(row=2, column=col_i, value=text)
        c.fill = make_fill(BG["blue"]); c.font = make_font(bold=True, color="FFFFFF")
        c.alignment = make_align("center"); c.border = make_border()
    ws.row_dimensions[2].height = 22

    # last year column in P&L and Cash Flow (both use the same A=label, B=yr0, ... layout)
    lc = get_column_letter(2 + n)

    kpi_items = [
        ("Доходы, руб.",           "P&L: Доходы",
         "='P&L'!B3",              f"='P&L'!{lc}3",             NUM),
        ("Расходы, руб.",          "P&L: Расходы",
         "='P&L'!B4",              f"='P&L'!{lc}4",             NUM),
        ("Чистый результат, руб.", "Доходы − Расходы",
         "='P&L'!B5",              f"='P&L'!{lc}5",             NUM),
        ("EBITDA, руб.",           "Рез-т + Амортизация",
         "='P&L'!B6",              f"='P&L'!{lc}6",             NUM),
        ("Рентабельность",         "Чистый рез-т / Доходы",
         "='P&L'!B7",              f"='P&L'!{lc}7",             PCT2),
        ("Рентабельность EBITDA",  "EBITDA / Доходы",
         "='P&L'!B8",              f"='P&L'!{lc}8",             PCT2),
        (
            "CAGR доходов",
            f"(Итог/База)^(1/{n})−1",
            "—",
            f"=IF('P&L'!B3=0,0,('P&L'!{lc}3/'P&L'!B3)^(1/{n})-1)",
            PCT2,
        ),
        ("Расходы / Доходы",       "Расходы / Доходы",
         "=IF('P&L'!B3=0,0,'P&L'!B4/'P&L'!B3)",
         f"=IF('P&L'!{lc}3=0,0,'P&L'!{lc}4/'P&L'!{lc}3)",     PCT),
        ("Операционный CF, руб.",  "Опер. деятельность",
         "='Денежный поток'!B5",   f"='Денежный поток'!{lc}5",  NUM),
        ("Чистый CF, руб.",        "Опер. + Инвест.",
         "='Денежный поток'!B8",   f"='Денежный поток'!{lc}8",  NUM),
        ("Накопленный CF, руб.",   "Нарастающим итогом",
         "='Денежный поток'!B9",   f"='Денежный поток'!{lc}9",  NUM),
        ("Доход на студента, руб.","Доходы / студентов",
         f"='P&L'!B3/{cfg['students']}",
         f"='P&L'!{lc}3/{cfg['students']}",                     NUM),
    ]

    for ri, (name, desc, base_f, last_f, num_fmt) in enumerate(kpi_items):
        r      = 3 + ri
        bg_key = "row" if ri % 2 == 0 else "white"

        data_cell(ws, r, 1, name, bold=True, bg=bg_key)
        data_cell(ws, r, 2, desc, bg=bg_key, fg="6B7280")

        for col_i, val in [(3, base_f), (4, last_f)]:
            c = ws.cell(row=r, column=col_i, value=val)
            c.number_format = num_fmt
            c.fill      = make_fill(BG["l_blue"] if col_i == 4 else BG[bg_key])
            c.font      = make_font(bold=(col_i == 4))
            c.border    = make_border()
            c.alignment = make_align("right")
        ws.row_dimensions[r].height = 20
