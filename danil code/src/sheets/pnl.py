"""
P&L row index reference (for cross-sheet formulas in other modules):
  3 = Доходы,  4 = Расходы,  5 = Чистый результат,
  6 = EBITDA,  7 = Рентабельность,  8 = Рентабельность EBITDA,
  9 = Прирост доходов г/г
"""
from openpyxl.utils import get_column_letter

from src.profiles import REVENUES, EXPENSES, DA_RATIO
from src.styles import (
    BG, NUM, PCT, PCT2,
    make_fill, make_font, make_border, make_align,
    sheet_title, apply_header_row, data_cell,
)


def build(wb, cfg: dict) -> None:
    ws = wb.create_sheet("P&L")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B3"

    years   = [cfg["base_year"] + i for i in range(cfg["years"] + 1)]
    n_years = len(years)
    rev_tr  = 3 + len(REVENUES[cfg["profile"]])
    exp_tr  = 3 + len(EXPENSES[cfg["profile"]])
    n_cols  = 1 + n_years

    ws.column_dimensions["A"].width = 30
    for i in range(n_years):
        ws.column_dimensions[get_column_letter(2 + i)].width = 16

    sheet_title(ws, f"P&L — {cfg['name'].upper()}", n_cols, "navy")

    ws.cell(row=2, column=1, value="Показатель")
    for i, y in enumerate(years):
        ws.cell(row=2, column=2 + i, value=str(y))
    apply_header_row(ws, 2, n_cols, "navy")

    def col(i: int) -> str:
        return get_column_letter(2 + i)

    da = DA_RATIO[cfg["profile"]]

    pnl_rows = [
        (
            "Доходы всего",
            [f"='Доходы'!{get_column_letter(3+i)}{rev_tr}" for i in range(n_years)],
            NUM, "l_green", True,
        ),
        (
            "Расходы всего",
            [f"='Расходы'!{get_column_letter(3+i)}{exp_tr}" for i in range(n_years)],
            NUM, "l_red", True,
        ),
        (
            "Чистый результат",
            [f"={col(i)}3-{col(i)}4" for i in range(n_years)],
            NUM, "l_blue", True,
        ),
        (
            "EBITDA",
            [f"={col(i)}5+{da}*{col(i)}3" for i in range(n_years)],
            NUM, "l_gray", False,
        ),
        (
            "Рентабельность",
            [f"=IF({col(i)}3=0,0,{col(i)}5/{col(i)}3)" for i in range(n_years)],
            PCT2, "l_gray", False,
        ),
        (
            "Рентабельность EBITDA",
            [f"=IF({col(i)}3=0,0,{col(i)}6/{col(i)}3)" for i in range(n_years)],
            PCT2, "l_gray", False,
        ),
        (
            "Прирост доходов г/г",
            ["—"] + [
                f"=IF({col(i-1)}3=0,0,{col(i)}3/{col(i-1)}3-1)"
                for i in range(1, n_years)
            ],
            PCT, "l_gray", False,
        ),
    ]

    for ri, (label, formulas, num_fmt, bg, bold) in enumerate(pnl_rows):
        r = 3 + ri
        data_cell(ws, r, 1, label, bold=bold, bg=bg)
        for ci, f_val in enumerate(formulas):
            c = ws.cell(row=r, column=2 + ci, value=f_val)
            c.number_format = num_fmt
            c.fill      = make_fill(BG[bg])
            c.font      = make_font(bold=bold)
            c.alignment = make_align("right")
            c.border    = make_border()
        ws.row_dimensions[r].height = 20
