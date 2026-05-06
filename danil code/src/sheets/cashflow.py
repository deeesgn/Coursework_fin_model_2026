"""
Cash Flow row index reference (for cross-sheet formulas in other modules):
  3 = Чистый результат,  4 = Амортизация,
  5 = Операционный CF,   6 = Капитальные затраты,
  7 = Инвестиционный CF, 8 = Чистый CF, 9 = Накопленный CF
"""
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter

from src.profiles import DA_RATIO, CAPEX_RATIO
from src.styles import (
    BG, NUM,
    make_fill, make_font, make_border, make_align,
    sheet_title, apply_header_row, data_cell,
)


def build(wb, cfg: dict) -> None:
    ws = wb.create_sheet("Денежный поток")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B3"

    years   = [cfg["base_year"] + i for i in range(cfg["years"] + 1)]
    n_years = len(years)
    n_cols  = 1 + n_years

    ws.column_dimensions["A"].width = 32
    for i in range(n_years):
        ws.column_dimensions[get_column_letter(2 + i)].width = 16

    sheet_title(ws, f"ДЕНЕЖНЫЙ ПОТОК — {cfg['name'].upper()}", n_cols, "teal")

    ws.cell(row=2, column=1, value="Показатель")
    for i, y in enumerate(years):
        ws.cell(row=2, column=2 + i, value=str(y))
    apply_header_row(ws, 2, n_cols, "teal")

    da_r    = DA_RATIO[cfg["profile"]]
    capex_r = CAPEX_RATIO[cfg["profile"]]

    def col(i: int) -> str:
        return get_column_letter(2 + i)

    cf_rows = [
        (
            "Чистый результат",
            [f"='P&L'!{col(i)}5" for i in range(n_years)],
            NUM, "l_gray", False,
        ),
        (
            "Амортизация (+)",
            [f"={da_r}*'P&L'!{col(i)}3" for i in range(n_years)],
            NUM, "l_gray", False,
        ),
        (
            "Операционный CF",
            [f"={col(i)}3+{col(i)}4" for i in range(n_years)],
            NUM, "l_teal", True,
        ),
        (
            "Капитальные затраты (-)",
            [f"=-{capex_r}*'P&L'!{col(i)}3" for i in range(n_years)],
            NUM, "l_red", False,
        ),
        (
            "Инвестиционный CF",
            [f"={col(i)}6" for i in range(n_years)],
            NUM, "l_red", True,
        ),
        (
            "Чистый денежный поток",
            [f"={col(i)}5+{col(i)}7" for i in range(n_years)],
            NUM, "l_blue", True,
        ),
        (
            "Накопленный CF",
            [f"=SUM({col(0)}8:{col(i)}8)" for i in range(n_years)],
            NUM, "l_blue", False,
        ),
    ]

    for ri, (label, formulas, num_fmt, bg, bold) in enumerate(cf_rows):
        r = 3 + ri
        data_cell(ws, r, 1, label, bold=bold, bg=bg)
        for ci, f_val in enumerate(formulas):
            c = ws.cell(row=r, column=2 + ci, value=f_val)
            c.number_format = num_fmt
            c.fill      = make_fill(BG[bg])
            c.font      = make_font(bold=bold)
            c.alignment = make_align("right")
            c.border    = make_border()
        ws.row_dimensions[r].height = 20

    # Жирная нижняя граница после Операционного CF (row 5) и Инвестиционного CF (row 7)
    medium = Side(style="medium", color="0D9488")
    for sep_row in (5, 7):
        for col_i in range(1, n_cols + 1):
            c = ws.cell(row=sep_row, column=col_i)
            b = c.border
            c.border = Border(left=b.left, right=b.right, top=b.top, bottom=medium)
