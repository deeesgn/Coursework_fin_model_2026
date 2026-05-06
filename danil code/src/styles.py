from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BG = {
    "navy":   "1E3A5F", "blue":    "2563EB",
    "green":  "059669", "red":     "DC2626",
    "teal":   "0D9488", "purple":  "7C3AED",
    "l_blue": "DBEAFE", "l_green": "D1FAE5",
    "l_red":  "FEE2E2", "l_gray":  "F3F4F6",
    "l_teal": "CCFBF1",
    "white":  "FFFFFF", "row":     "F9FAFB",
}

NUM  = "#,##0"
PCT  = "0.0%"
PCT2 = "0.00%"


def make_fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def make_font(bold: bool = False, color: str = "111827",
              size: int = 10, italic: bool = False) -> Font:
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")


def make_border() -> Border:
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)


def make_align(h: str = "left", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)


def set_col_widths(ws, widths: dict) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def sheet_title(ws, text: str, n_cols: int, color: str = "navy") -> None:
    last_col = get_column_letter(n_cols)
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"].value     = text
    ws["A1"].fill      = make_fill(BG[color])
    ws["A1"].font      = make_font(bold=True, size=11, color="FFFFFF")
    ws["A1"].alignment = make_align("center")
    ws.row_dimensions[1].height = 26


def apply_header_row(ws, row: int, n_cols: int, color: str = "blue") -> None:
    for col in range(1, n_cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill      = make_fill(BG[color])
        c.font      = make_font(bold=True, color="FFFFFF")
        c.alignment = make_align("center")
        c.border    = make_border()


def data_cell(ws, row: int, col: int, value,
              num_fmt: str | None = None,
              bold: bool = False, bg: str = "white", fg: str = "111827"):
    c = ws.cell(row=row, column=col, value=value)
    c.fill      = make_fill(BG[bg])
    c.font      = make_font(bold=bold, color=fg)
    c.alignment = make_align("right" if col > 1 else "left")
    c.border    = make_border()
    if num_fmt:
        c.number_format = num_fmt
    return c


def apply_total_row(ws, row: int, n_cols: int, bg: str = "l_blue") -> None:
    for col in range(1, n_cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill      = make_fill(BG[bg])
        c.font      = make_font(bold=True)
        c.alignment = make_align("right" if col > 1 else "left")
        c.border    = make_border()
