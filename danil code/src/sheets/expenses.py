from openpyxl.utils import get_column_letter

from src.profiles import EXPENSES
from src.styles import (
    BG, NUM, PCT,
    make_fill, make_font, make_border, make_align,
    sheet_title, apply_header_row, apply_total_row, data_cell,
)


def build(wb, cfg: dict) -> None:
    ws = wb.create_sheet("Расходы")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C3"

    years   = [cfg["base_year"] + i for i in range(cfg["years"] + 1)]
    n_years = len(years)
    rows    = EXPENSES[cfg["profile"]]
    n_cols  = 2 + n_years

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 11
    for i in range(n_years):
        ws.column_dimensions[get_column_letter(3 + i)].width = 16

    sheet_title(ws, f"РАСХОДЫ — {cfg['name'].upper()}", n_cols, "red")

    ws.cell(row=2, column=1, value="Статья расходов")
    ws.cell(row=2, column=2, value="Рост %")
    for i, y in enumerate(years):
        ws.cell(row=2, column=3 + i, value=str(y))
    apply_header_row(ws, 2, n_cols, "red")

    for ri, (label, base, growth) in enumerate(rows):
        r  = 3 + ri
        bg = "row" if ri % 2 == 0 else "white"

        data_cell(ws, r, 1, label,  bg=bg)
        data_cell(ws, r, 2, growth, PCT, bg=bg, fg="D97706")
        data_cell(ws, r, 3, base,   NUM, bold=True, bg=bg)

        for ci in range(1, n_years):
            prev = get_column_letter(3 + ci - 1)
            data_cell(ws, r, 3 + ci, f"={prev}{r}*(1+B{r})", NUM, bg=bg)

    tr = 3 + len(rows)
    ws.cell(row=tr, column=1, value="ИТОГО РАСХОДОВ")
    apply_total_row(ws, tr, n_cols, "l_red")
    for ci in range(n_years):
        cl = get_column_letter(3 + ci)
        c  = ws.cell(row=tr, column=3 + ci, value=f"=SUM({cl}3:{cl}{tr-1})")
        c.number_format = NUM; c.font = make_font(bold=True)
        c.border = make_border(); c.alignment = make_align("right")
        c.fill = make_fill(BG["l_red"])
    ws.row_dimensions[tr].height = 22
