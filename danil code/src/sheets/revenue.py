from openpyxl.utils import get_column_letter

from src.profiles import REVENUES
from src.styles import (
    BG, NUM, PCT, PCT2,
    make_fill, make_font, make_border, make_align,
    sheet_title, apply_header_row, apply_total_row, data_cell,
)


def build(wb, cfg: dict) -> None:
    ws = wb.create_sheet("Доходы")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C3"

    years   = [cfg["base_year"] + i for i in range(cfg["years"] + 1)]
    n_years = len(years)
    rows    = REVENUES[cfg["profile"]]
    n_cols  = 2 + n_years + 1  # label | рост % | годы | CAGR

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 11
    for i in range(n_years):
        ws.column_dimensions[get_column_letter(3 + i)].width = 16
    ws.column_dimensions[get_column_letter(3 + n_years)].width = 12

    sheet_title(ws, f"ДОХОДЫ — {cfg['name'].upper()}", n_cols, "green")

    ws.cell(row=2, column=1, value="Статья доходов")
    ws.cell(row=2, column=2, value="Рост %")
    for i, y in enumerate(years):
        ws.cell(row=2, column=3 + i, value=str(y))
    ws.cell(row=2, column=3 + n_years, value="CAGR")
    apply_header_row(ws, 2, n_cols, "green")

    for ri, (label, base, growth) in enumerate(rows):
        r  = 3 + ri
        bg = "row" if ri % 2 == 0 else "white"

        data_cell(ws, r, 1, label,  bg=bg)
        data_cell(ws, r, 2, growth, PCT, bg=bg, fg="059669")
        data_cell(ws, r, 3, base,   NUM, bold=True, bg=bg)

        for ci in range(1, n_years):
            prev = get_column_letter(3 + ci - 1)
            data_cell(ws, r, 3 + ci, f"={prev}{r}*(1+B{r})", NUM, bg=bg)

        fc   = get_column_letter(3)
        lc   = get_column_letter(3 + n_years - 1)
        n    = n_years - 1
        cagr = f"=IF({fc}{r}=0,\"\",({lc}{r}/{fc}{r})^(1/{n})-1)"
        data_cell(ws, r, 3 + n_years, cagr, PCT2, bg=bg, fg="7C3AED")

    tr = 3 + len(rows)
    ws.cell(row=tr, column=1, value="ИТОГО ДОХОДОВ")
    apply_total_row(ws, tr, n_cols, "l_green")
    for ci in range(n_years):
        cl = get_column_letter(3 + ci)
        c  = ws.cell(row=tr, column=3 + ci, value=f"=SUM({cl}3:{cl}{tr-1})")
        c.number_format = NUM; c.font = make_font(bold=True)
        c.border = make_border(); c.alignment = make_align("right")
        c.fill = make_fill(BG["l_green"])
    ws.row_dimensions[tr].height = 22
