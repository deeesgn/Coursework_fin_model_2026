from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

from src.styles import BG, NUM, make_fill, make_font, make_border, make_align, set_col_widths


def build(wb, cfg: dict) -> None:
    ws = wb.create_sheet("График")
    ws.sheet_view.showGridLines = False

    years   = [cfg["base_year"] + i for i in range(cfg["years"] + 1)]
    n_years = len(years)

    # Вспомогательная таблица (ссылки на P&L) — источник данных для графика
    for ci, header in enumerate(["Год", "Доходы", "Расходы", "Чистый результат"], start=1):
        c = ws.cell(row=1, column=ci, value=header)
        c.fill = make_fill(BG["navy"]); c.font = make_font(bold=True, color="FFFFFF")
        c.alignment = make_align("center"); c.border = make_border()

    for i, year in enumerate(years):
        r      = 2 + i
        pl_col = get_column_letter(2 + i)
        ws.cell(row=r, column=1, value=year)
        for col_i, pnl_row in [(2, 3), (3, 4), (4, 5)]:
            c = ws.cell(row=r, column=col_i, value=f"='P&L'!{pl_col}{pnl_row}")
            c.number_format = NUM; c.border = make_border()

    set_col_widths(ws, {"A": 10, "B": 18, "C": 18, "D": 18})

    chart = BarChart()
    chart.type      = "col"
    chart.grouping  = "clustered"
    chart.title     = f"Доходы vs Расходы — {cfg['name']}"
    chart.y_axis.title = "Руб."
    chart.x_axis.title = "Год"
    chart.style     = 10
    chart.width     = 22
    chart.height    = 14

    for col_i in (2, 3, 4):
        series_data = Reference(ws, min_col=col_i, min_row=1, max_row=1 + n_years)
        chart.add_data(series_data, titles_from_data=True)

    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=1 + n_years))
    ws.add_chart(chart, "F2")
